import requests
import time
from io import BytesIO
from openpyxl import load_workbook
import re
import os
from flask import Flask
import threading
from dotenv import load_dotenv
# ================= CONFIG =================
app = Flask(__name__)

load_dotenv()
TENANT_ID = os.getenv("TENANT_ID")
CLIENT_ID = os.getenv("CLIENT_ID")
CLIENT_SECRET = os.getenv("CLIENT_SECRET")

DRIVE_ID = os.getenv("DRIVE_ID")
ITEM_ID = os.getenv("ITEM_ID")
REFRESH_TOKEN = os.getenv('REFRESH_TOKEN')
POLL_INTERVAL = 60  # seconds
# ==========================================
def refresh_access_token()-> str:
    global ACCESS_TOKEN
    url = "https://accounts.zoho.com/oauth/v2/token"
    params = {
        "refresh_token": REFRESH_TOKEN,
        "client_id": CLIENT_ID,
        "client_secret": CLIENT_SECRET,
        "grant_type": "refresh_token"
    }
    resp = requests.post(url, params=params).json()
    ACCESS_TOKEN = resp["access_token"]
    return ACCESS_TOKEN
def get_token():
    url = f"https://login.microsoftonline.com/{TENANT_ID}/oauth2/v2.0/token"
    data = {
        "client_id": CLIENT_ID,
        "client_secret": CLIENT_SECRET,
        "grant_type": "client_credentials",
        "scope": "https://graph.microsoft.com/.default"
    }

    resp = requests.post(url, data=data).json()

    if "access_token" not in resp:
        raise Exception(f"Token error: {resp}")

    return resp["access_token"]


def download_excel(headers):
    url = f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}/items/{ITEM_ID}/content"
    return requests.get(url, headers=headers).content


def upload_excel(headers, data):
    url = f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}/items/{ITEM_ID}/content"
    requests.put(url, headers=headers, data=data)
ID_RE = re.compile(r"_(\d{4})$")

def assign_ids(file_bytes):
    wb = load_workbook(BytesIO(file_bytes))
    global_ws = wb["__GLOBAL__"]

    # 1️⃣ Collect all existing IDs
    used_ids = set()
    for ws in wb.worksheets:
        if ws.title == "__GLOBAL__":
            continue
        for row in ws.iter_rows(min_row=2):
            name = row[1].value
            if isinstance(name, str):
                m = ID_RE.search(name)
                if m:
                    used_ids.add(int(m.group(1)))

    # 2️⃣ Start assigning new IDs to missing ones
    last_id = max(used_ids) if used_ids else 0
    changed = False

    for ws in wb.worksheets:
        if ws.title == "__GLOBAL__":
            continue
        for row in ws.iter_rows(min_row=2):
            cell = row[1]
            name = cell.value
            if not isinstance(name, str) or not name.strip():
                continue

            m = ID_RE.search(name)
            if m:
                continue  # already has ID

            # Find the lowest available ID
            new_id = 1
            while new_id in used_ids:
                new_id += 1

            used_ids.add(new_id)
            cell.value = f"{name}_{new_id:04d}"
            changed = True
            last_id = max(last_id, new_id)

    # 3️⃣ Persist correct global counter
    if changed:
        global_ws["B1"].value = last_id
        out = BytesIO()
        wb.save(out)
        out.seek(0)
        return out.read(), last_id

    return None, last_id


def get_last_modified(headers):
    url = f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}/items/{ITEM_ID}"
    return requests.get(url, headers=headers).json()["lastModifiedDateTime"]


def main():
    token = get_token() or refresh_access_token()
    headers = {"Authorization": f"Bearer {token}"}

    last_seen = None

    print("📡 Watching SharePoint file for changes...")

    while True:
        try:
            modified = get_last_modified(headers)

            if modified != last_seen:
                print("🔄 File changed, processing...")
                file_bytes = download_excel(headers)

                result, last_id = assign_ids(file_bytes)

                if result:
                    upload_excel(headers, result)
                    print(f"✅ IDs assigned. Last ID = {last_id}")
                else:
                    print("ℹ No new projects found")

                last_seen = modified

        except Exception as e:
            print("⚠ Error:", e)

        time.sleep(POLL_INTERVAL)

threading.Thread(target=main, daemon=True).start()
@app.route("/")
def index():
    return "Excel watcher running ✅"

@app.route("/health")
def health():
    return {'health':'ok'}

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 8080))
    app.run(host="0.0.0.0", port=port)